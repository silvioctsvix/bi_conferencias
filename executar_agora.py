#!/usr/bin/env python
# coding: utf-8

import logging
import traceback
import obter_data_atualizacao as od
import os
import openpyxl
from datetime import datetime

# Configurar logging
logging.basicConfig(
    filename='log_execucao_manual.log', 
    level=logging.INFO, 
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s"
)

def criar_planilha_controle():
    """
    Cria um novo arquivo Excel de controle para a última atualização
    """
    print("Criando arquivo de controle de última atualização...")
    logging.info("Criando arquivo de controle de última atualização")
    
    # Definir caminhos
    diretorio = r"D:\\Silvio\\OneDrive\\Trabalho\\POWER BI\\AUDITORIA PROCESSO SS DIGITAL - NEW\\"
    arquivo = "UltimaAtualização.xlsx"
    caminho_completo = os.path.join(diretorio, arquivo)
    
    # Verificar e criar diretório se não existir
    if not os.path.exists(diretorio):
        try:
            os.makedirs(diretorio, exist_ok=True)
            logging.info(f"Diretório criado: {diretorio}")
        except Exception as e:
            logging.error(f"Erro ao criar diretório: {str(e)}")
            # Tentar diretório alternativo
            diretorio = r"D:\\Temp\\"
            caminho_completo = os.path.join(diretorio, arquivo)
            logging.info(f"Tentando diretório alternativo: {diretorio}")
            if not os.path.exists(diretorio):
                os.makedirs(diretorio, exist_ok=True)
    
    # Criar nova planilha
    try:
        wb = openpyxl.Workbook()
        
        # Renomear a aba ativa para 'Ultima'
        ws = wb.active
        ws.title = 'Ultima'
        
        # Adicionar cabeçalho
        ws['A1'] = 'Última Atualização'
        
        # Adicionar data atual como a última atualização
        data_atual = datetime.now().strftime('%d/%m/%Y')
        ws['A2'] = data_atual
        
        # Salvar arquivo
        wb.save(caminho_completo)
        print(f"Arquivo Excel criado com sucesso: {caminho_completo}")
        print(f"Data inicial configurada: {data_atual}")
        logging.info(f"Arquivo Excel criado com sucesso: {caminho_completo}")
        
        return True
    except Exception as e:
        msg = f"Erro ao criar arquivo Excel: {str(e)}"
        print(msg)
        logging.error(msg)
        logging.error(traceback.format_exc())
        return False

def verificar_arquivo_controle():
    """Verifica se o arquivo de controle existe e o cria se necessário"""
    diretorio = r"D:\\Silvio\\OneDrive\\Trabalho\\POWER BI\\AUDITORIA PROCESSO SS DIGITAL - NEW\\"
    arquivo = "UltimaAtualização.xlsx"
    caminho_completo = os.path.join(diretorio, arquivo)
    
    if not os.path.exists(caminho_completo):
        print(f"Arquivo de controle não encontrado: {caminho_completo}")
        print("Criando novo arquivo de controle...")
        return criar_planilha_controle()
    
    # Testar se o arquivo está corrompido
    try:
        openpyxl.load_workbook(caminho_completo)
        print("Arquivo de controle existente e válido")
        return True
    except Exception as e:
        print(f"Arquivo de controle existente mas corrompido: {str(e)}")
        print("Criando novo arquivo de controle...")
        # Tentar remover o arquivo corrompido
        try:
            os.remove(caminho_completo)
        except:
            pass
        return criar_planilha_controle()

def executar_atualizacao():
    """Executa a atualização diretamente sem agendamento"""
    try:
        print("Verificando arquivo de controle...")
        if not verificar_arquivo_controle():
            return False
            
        print("Iniciando atualização...")
        logging.info("Iniciando execução manual da atualização")
        od.obter_data()
        logging.info("Execução manual concluída com sucesso")
        print("Atualização concluída com sucesso!")
        return True
    except Exception as e:
        erro_msg = f"Erro na execução manual: {str(e)}"
        logging.error(erro_msg)
        logging.error(traceback.format_exc())
        print(f"ERRO: {erro_msg}")
        return False

if __name__ == "__main__":
    executar_atualizacao()