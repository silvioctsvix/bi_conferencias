#!/usr/bin/env python
# coding: utf-8

import openpyxl
from datetime import datetime
import os

def criar_planilha_controle():
    """
    Cria um novo arquivo Excel de controle para a última atualização
    """
    print("Criando arquivo de controle de última atualização...")
    
    # Definir caminhos
    diretorio = r"D:\\Silvio\\OneDrive\\Trabalho\\POWER BI\\AUDITORIA PROCESSO SS DIGITAL - NEW\\"
    arquivo = "UltimaAtualização.xlsx"
    caminho_completo = os.path.join(diretorio, arquivo)
    
    # Verificar e criar diretório se não existir
    if not os.path.exists(diretorio):
        try:
            os.makedirs(diretorio, exist_ok=True)
            print(f"Diretório criado: {diretorio}")
        except Exception as e:
            print(f"Erro ao criar diretório: {str(e)}")
            # Tentar diretório alternativo
            diretorio = r"D:\\Temp\\"
            caminho_completo = os.path.join(diretorio, arquivo)
            print(f"Tentando diretório alternativo: {diretorio}")
            if not os.path.exists(diretorio):
                os.makedirs(diretorio, exist_ok=True)
    
    # Criar nova planilha
    try:
        wb = openpyxl.Workbook()
        
        # Renomear a aba ativa para 'Ultima'
        ws = wb.active
        ws.title = 'Ultima'
        
        # Adicionar cabeçalho
        ws['A1'] = 'Última Atualização'
        
        # Adicionar data atual como a última atualização
        data_atual = datetime.now().strftime('%d/%m/%Y')
        ws['A2'] = data_atual
        
        # Salvar arquivo
        wb.save(caminho_completo)
        print(f"Arquivo Excel criado com sucesso: {caminho_completo}")
        print(f"Data inicial configurada: {data_atual}")
        
        return True
    except Exception as e:
        print(f"Erro ao criar arquivo Excel: {str(e)}")
        return False

if __name__ == "__main__":
    criar_planilha_controle() 