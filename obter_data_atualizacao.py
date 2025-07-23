#!/usr/bin/env python
# coding: utf-8
#%load_ext autoreload
#%autoreload 2
import pandas as pd
from datetime import datetime
from datetime import date, timedelta
import openpyxl as opp
import GeraCSV_MovimentoDoDia as gera
import tempfile
import os
import shutil
import logging
import sys
import traceback
import stat

# Configurar logging
logging.basicConfig(
    filename='log_obter_data.log',
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)

def validar_data(valor):
    try:
        # Tente converter o valor em uma data usando o formato desejado
        #datetime.strptime(valor, '%Y-%m-%d')  # Formato: AAAA-MM-DD
        datetime.strptime(valor, '%d/%m/%Y')  # Formato: AAAA-MM-DD
        return True
    except ValueError:
        return False

def verificar_diretorio(caminho_diretorio):
    """Verifica se o diretório existe e pode ser escrito"""
    try:
        if not os.path.exists(caminho_diretorio):
            logging.info(f"Diretório {caminho_diretorio} não existe, tentando criar.")
            os.makedirs(caminho_diretorio, exist_ok=True)
            logging.info(f"Diretório {caminho_diretorio} criado com sucesso.")
        
        # Testar permissões de escrita
        if not os.access(caminho_diretorio, os.W_OK):
            logging.warning(f"Diretório {caminho_diretorio} não tem permissão de escrita.")
            return False
        
        # Tenta criar um arquivo de teste para verificar permissões
        teste_arquivo = os.path.join(caminho_diretorio, "teste_permissao.tmp")
        try:
            with open(teste_arquivo, 'w') as f:
                f.write('teste')
            os.remove(teste_arquivo)
            logging.info(f"Diretório {caminho_diretorio} tem permissão de escrita.")
            return True
        except Exception as e:
            logging.error(f"Erro ao testar permissões no diretório {caminho_diretorio}: {str(e)}")
            return False
    except Exception as e:
        logging.error(f"Erro ao verificar diretório {caminho_diretorio}: {str(e)}")
        return False

def verificar_remover_readonly(caminho_arquivo):
    """Verifica se o arquivo está em modo somente leitura e remove esse atributo se necessário"""
    if os.path.exists(caminho_arquivo):
        # Verificar os atributos do arquivo
        try:
            atributos = os.stat(caminho_arquivo).st_mode
            is_readonly = not (atributos & stat.S_IWRITE)
            
            if is_readonly:
                logging.info(f"Arquivo {caminho_arquivo} está em modo somente leitura. Tentando remover esse atributo.")
                # Remover atributo de somente leitura
                os.chmod(caminho_arquivo, atributos | stat.S_IWRITE)
                logging.info("Atributo somente leitura removido com sucesso.")
                return True
            else:
                logging.info(f"Arquivo {caminho_arquivo} não está em modo somente leitura.")
                return False
        except Exception as e:
            logging.error(f"Erro ao verificar/modificar atributos do arquivo: {str(e)}")
            logging.error(traceback.format_exc())
            return False
    else:
        logging.warning(f"Arquivo {caminho_arquivo} não existe.")
        return False

def salvar_planilha_seguro(pasta_trabalho, caminho_destino):
    """Salva a planilha em um local temporário e depois move para o destino final"""
    logging.info(f"Iniciando salvamento seguro para {caminho_destino}")
    
    # Verificar e remover atributo readonly se necessário
    verificar_remover_readonly(caminho_destino)
    
    # Verificar diretório de destino
    diretorio_destino = os.path.dirname(caminho_destino)
    if not verificar_diretorio(diretorio_destino):
        logging.error(f"Diretório de destino {diretorio_destino} não tem permissões adequadas.")
        # Tentar usar um diretório alternativo
        diretorio_destino = r"D:\Temp"
        caminho_destino = os.path.join(diretorio_destino, os.path.basename(caminho_destino))
        logging.info(f"Tentando usar diretório alternativo: {diretorio_destino}")
        verificar_diretorio(diretorio_destino)
    
    # Criar diretório temporário local se não existir
    temp_dir = r"D:\Temp"
    if not os.path.exists(temp_dir):
        os.makedirs(temp_dir)
        logging.info(f"Diretório temporário criado: {temp_dir}")
    
    # Nome de arquivo temporário
    nome_arquivo = os.path.basename(caminho_destino)
    caminho_temp = os.path.join(temp_dir, nome_arquivo)
    logging.info(f"Arquivo temporário: {caminho_temp}")
    
    # Configurar diretório temporário para o OpenPyXL
    temp_old = tempfile.tempdir
    tempfile.tempdir = temp_dir
    logging.info(f"Diretório temporário configurado para: {tempfile.tempdir}")
    
    try:
        # Salvar no local temporário
        logging.info("Tentando salvar no local temporário...")
        pasta_trabalho.save(caminho_temp)
        logging.info("Arquivo salvo com sucesso no local temporário")
        
        # Copiar para o destino final
        logging.info(f"Copiando para destino final: {caminho_destino}")
        shutil.copy2(caminho_temp, caminho_destino)
        logging.info("Cópia concluída com sucesso")
        
        # Remover arquivo temporário
        os.remove(caminho_temp)
        logging.info("Arquivo temporário removido")
        
        return True
    except Exception as e:
        logging.error(f"Erro ao salvar planilha: {str(e)}")
        logging.error(traceback.format_exc())
        
        # Tentar salvar diretamente no destino se a cópia falhar
        try:
            logging.info("Tentando salvar diretamente no destino final...")
            pasta_trabalho.save(caminho_destino)
            logging.info("Arquivo salvo com sucesso diretamente no destino")
            return True
        except Exception as e2:
            logging.error(f"Erro ao salvar diretamente no destino: {str(e2)}")
            return False
    finally:
        # Restaurar configuração original
        tempfile.tempdir = temp_old
        logging.info("Configuração de tempdir restaurada")

def obter_data():
    logging.info("Iniciando obter_data()")
    diretorio = r"D:\\Silvio\\OneDrive\\Trabalho\\POWER BI\\AUDITORIA PROCESSO SS DIGITAL - NEW\\"
    arquivo   = "UltimaAtualização.xlsx"
    planilha  = 'Ultima'
    linha     = 2
    coluna    = 1
    data_ultima_atualização = date(1900, 1, 1)
    lista_de_datas = []
    delta = timedelta(days=1) #intervalo de 1 dia
    resultado = ''
    tipo_dado = 's'  # "s" para string, "n" para número, "d" para data, etc.
    tipo_movimento = ''

    # Verificar diretório
    logging.info(f"Verificando permissões do diretório: {diretorio}")
    verificar_diretorio(diretorio)

    caminho_completo = os.path.join(diretorio, arquivo)
    logging.info(f"Verificando permissões do arquivo: {caminho_completo}")
    verificar_remover_readonly(caminho_completo)

    logging.info(f"Tentando abrir planilha: {caminho_completo}")
    try:
        pasta_trabalho = opp.load_workbook(caminho_completo)
        plan = pasta_trabalho[planilha]
        ultima_atualização = plan.cell(row=linha, column=coluna).value
        logging.info(f"Última atualização lida da planilha: {ultima_atualização}")
        #ultima_atualização = ultima_atualização[0:10]
    except Exception as erro:
        logging.error(f"Erro ao abrir planilha: {str(erro)}")
        logging.error(traceback.format_exc())
        raise ValueError("Erro ao abrir planilha UltimaAtualização.xlsx!: " + str(erro))

    if (validar_data(ultima_atualização) == True):
        data_ultima_atualização = datetime.strptime(ultima_atualização, '%d/%m/%Y').date()
        logging.info(f"Data última atualização validada: {data_ultima_atualização}")
    else:
        msg = 'Data última atualização inválida!'
        logging.error(msg)
        print(msg)
        raise ValueError(msg)

    #Obtem a lita de datas a serem atualizadas
    data_inicial = data_ultima_atualização
    data_atual = datetime.now().date()
    logging.info(f"Data atual: {data_atual}")
    
    while data_inicial <= data_atual:
        lista_de_datas.append(data_inicial)
        data_inicial += delta
    
    logging.info(f"Datas a processar: {lista_de_datas}")

    for data_a_atualizar in lista_de_datas:
        logging.info(f"Processando data: {data_a_atualizar}")
        if data_atual == data_a_atualizar:
            tipo_movimento = 'incremental'
        else:
            tipo_movimento = 'normal'
        
        logging.info(f"Tipo de movimento: {tipo_movimento}")
        resultado = gera.GeraCSVMovimentoDia(data_a_atualizar.strftime("%d/%m/%Y"), tipo_movimento)
        
        if resultado == None:
            logging.info("Geração de CSV concluída com sucesso")
            plan.cell(row=linha, column=coluna, value=data_a_atualizar.strftime("%d/%m/%Y"))
            logging.info(f"Célula atualizada com data: {data_a_atualizar.strftime('%d/%m/%Y')}")
            
            # Verificar novamente as permissões antes de salvar
            verificar_remover_readonly(caminho_completo)
            
            # Usar função de salvamento seguro
            sucesso = salvar_planilha_seguro(pasta_trabalho, caminho_completo)
            if not sucesso:
                logging.error("Falha ao salvar planilha de forma segura")
                
                # Tentar salvar em caminho alternativo se falhar
                caminho_alternativo = rf"D:\Temp\{arquivo}"
                logging.info(f"Tentando salvar em caminho alternativo: {caminho_alternativo}")
                sucesso_alt = salvar_planilha_seguro(pasta_trabalho, caminho_alternativo)
                if sucesso_alt:
                    logging.info(f"Arquivo salvo com sucesso no caminho alternativo: {caminho_alternativo}")
                    print(f"Arquivo salvo em caminho alternativo: {caminho_alternativo}")
        else: 
            logging.error(f"Erro na geração de CSV: {resultado}")
            print(resultado)
    
    logging.info("Função obter_data concluída com sucesso")
